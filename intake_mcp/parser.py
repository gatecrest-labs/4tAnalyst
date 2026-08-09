"""
Firewall Request Spreadsheet Parser

Handles the standard multi-tab .xlsx firewall request form with four tabs:
  - FW Rules       : Rule#, Add/Remove, Source IP/Host, Dest IP/Host, Port#, Protocol,
                     Application, Business Justification
  - Group Request  : Group, New Group Membership Should Be, Action Required
  - IP Block/Allow : IP address, Business Justification  (title row 1, header row 3)
  - VPN            : Local + Remote side tunnel params, ISAKMP/IPSec params,
                     shared networks/host routes

Tab names are matched case-insensitively and by keyword so minor naming variations
in different versions of the template are handled gracefully.

The normalised output is a FirewallRequest dataclass — the same structure regardless
of whether data came from a spreadsheet or manual entry via parse_manual_entry().
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class FWRule:
    rule_number: str = ""
    operation: str = ""          # Add | Remove
    source: str = ""             # IP, CIDR, or hostname
    destination: str = ""
    port: str = ""
    protocol: str = ""           # TCP | UDP | Any | N/A
    application: str = ""
    justification: str = ""
    missing_fields: list[str] = field(default_factory=list)


@dataclass
class GroupChange:
    group_name: str = ""
    new_membership: str = ""
    action: str = ""             # Add | Remove | Replace


@dataclass
class IPBlockAllow:
    ip_address: str = ""
    operation: str = ""          # Block | Allow  (from tab title row)
    justification: str = ""


@dataclass
class VPNConfig:
    # Local side
    local_device: str = ""
    local_tunnel_endpoint: str = ""
    local_ike_encryption: str = ""
    local_auth_method: str = ""
    local_dh_group: str = ""
    local_sa_lifetime_p1: str = ""
    local_hash_p1: str = ""
    local_ipsec_encryption: str = ""
    local_hash_p2: str = ""
    local_sa_lifetime_p2: str = ""
    local_pfs: str = ""
    local_networks: list[str] = field(default_factory=list)
    # Remote side
    remote_device: str = ""
    remote_tunnel_endpoint: str = ""
    remote_ike_encryption: str = ""
    remote_auth_method: str = ""
    remote_dh_group: str = ""
    remote_sa_lifetime_p1: str = ""
    remote_hash_p1: str = ""
    remote_ipsec_encryption: str = ""
    remote_hash_p2: str = ""
    remote_sa_lifetime_p2: str = ""
    remote_pfs: str = ""
    remote_networks: list[str] = field(default_factory=list)


@dataclass
class FirewallRequest:
    # Request metadata
    reference_id: str = ""       # RITM, CHG, or any identifier the engineer provides
    rule_owner: str = ""
    business_contact: str = ""
    business_area: str = ""
    nuclear_environment: bool = False
    designation: str = ""        # e.g. "OT", "CIP", "IT"
    estimated_completion: str = ""
    source_file: str = ""        # original filename for traceability

    # Per-tab content
    fw_rules: list[FWRule] = field(default_factory=list)
    group_changes: list[GroupChange] = field(default_factory=list)
    ip_block_allow: list[IPBlockAllow] = field(default_factory=list)
    vpn_config: VPNConfig | None = None

    # Validation
    missing_fields: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    def has_fw_rules(self) -> bool:
        return bool(self.fw_rules)

    def has_vpn(self) -> bool:
        return self.vpn_config is not None and bool(
            self.vpn_config.local_tunnel_endpoint or self.vpn_config.remote_tunnel_endpoint
        )

    def has_group_changes(self) -> bool:
        return bool(self.group_changes)

    def has_ip_block_allow(self) -> bool:
        return bool(self.ip_block_allow)


# ---------------------------------------------------------------------------
# Spreadsheet parser
# ---------------------------------------------------------------------------

def parse_spreadsheet(file_path: str | Path) -> FirewallRequest:
    """
    Parse a firewall request .xlsx file.

    Raises FileNotFoundError if the file doesn't exist.
    Raises ImportError if openpyxl is not installed.
    Returns a FirewallRequest; never raises on malformed data — problems go
    into FirewallRequest.warnings instead.
    """
    if not _OPENPYXL:
        raise ImportError(
            "openpyxl is required to parse spreadsheets. "
            "Install it with: uv pip install openpyxl"
        )

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    req = FirewallRequest(source_file=path.name)

    for sheet in wb.worksheets:
        name_lower = sheet.title.lower()
        if _tab_matches(name_lower, ["fw rule", "firewall rule", "fw rules", "firewall rules", "rules"]):
            _parse_fw_rules_tab(sheet, req)
        elif _tab_matches(name_lower, ["group"]):
            _parse_group_tab(sheet, req)
        elif _tab_matches(name_lower, ["ip block", "block", "allow", "url"]):
            _parse_ip_block_allow_tab(sheet, req)
        elif _tab_matches(name_lower, ["vpn"]):
            _parse_vpn_tab(sheet, req)

    _validate(req)
    return req


def _tab_matches(name: str, keywords: list[str]) -> bool:
    return any(kw in name for kw in keywords)


# ---------------------------------------------------------------------------
# FW Rules tab
# ---------------------------------------------------------------------------

def _parse_fw_rules_tab(ws: Worksheet, req: FirewallRequest) -> None:
    """
    Header row contains: Rule #, Add/Remove, Source IP/Host Name,
    Destination IP/Host Name, Port #, Protocol, Application, Justification.
    Rows with no source AND no destination are skipped.
    """
    header_row: dict[str, int] = {}  # normalised col name → 0-based col index

    for row in ws.iter_rows(values_only=True):
        # Find the header row (contains "source" or "add")
        if not header_row:
            row_lower = [str(c).lower().strip() if c else "" for c in row]
            if any(k in " ".join(row_lower) for k in ["source", "add / remove", "add/remove"]):
                for i, cell in enumerate(row_lower):
                    if "rule" in cell and "#" in cell:
                        header_row["rule_number"] = i
                    elif "add" in cell and ("remove" in cell or "/" in cell):
                        header_row["operation"] = i
                    elif "source" in cell:
                        header_row["source"] = i
                    elif "dest" in cell:
                        header_row["destination"] = i
                    elif "port" in cell and "application" not in cell and "using" not in cell:
                        header_row["port"] = i
                    elif "protocol" in cell:
                        header_row["protocol"] = i
                    elif "application" in cell:
                        header_row["application"] = i
                    elif "justif" in cell or "comment" in cell or "business" in cell:
                        header_row["justification"] = i
            continue

        rule = FWRule(
            rule_number=_cell(row, header_row.get("rule_number")),
            operation=_cell(row, header_row.get("operation")) or "Add",
            source=_cell(row, header_row.get("source")),
            destination=_cell(row, header_row.get("destination")),
            port=_cell(row, header_row.get("port")),
            protocol=_cell(row, header_row.get("protocol")) or "N/A",
            application=_cell(row, header_row.get("application")),
            justification=_cell(row, header_row.get("justification")),
        )

        # Skip completely empty rows
        if not rule.source and not rule.destination:
            continue

        # Flag missing required fields per rule
        if not rule.source:
            rule.missing_fields.append("source")
        if not rule.destination:
            rule.missing_fields.append("destination")
        if not rule.port:
            rule.missing_fields.append("port")
        if not rule.justification:
            rule.missing_fields.append("justification")

        req.fw_rules.append(rule)


# ---------------------------------------------------------------------------
# Group Request tab
# ---------------------------------------------------------------------------

def _parse_group_tab(ws: Worksheet, req: FirewallRequest) -> None:
    """
    Columns: Group (A), New Group Membership Should Be (C), Action Required (D).
    Header is row 1; data starts row 2.
    """
    header_found = False
    col_map: dict[str, int] = {}

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            row_lower = [str(c).lower().strip() if c else "" for c in row]
            if any("group" in c for c in row_lower):
                for i, c in enumerate(row_lower):
                    if c == "group":
                        col_map["group"] = i
                    elif "membership" in c or "new group" in c:
                        col_map["membership"] = i
                    elif "action" in c:
                        col_map["action"] = i
                header_found = True
            continue

        group = _cell(row, col_map.get("group"))
        membership = _cell(row, col_map.get("membership"))
        action = _cell(row, col_map.get("action"))

        if not group and not membership:
            continue

        req.group_changes.append(GroupChange(
            group_name=group,
            new_membership=membership,
            action=action,
        ))


# ---------------------------------------------------------------------------
# IP Block/Allow tab
# ---------------------------------------------------------------------------

def _parse_ip_block_allow_tab(ws: Worksheet, req: FirewallRequest) -> None:
    """
    Row 1: title containing "IP address block or allow"
    Row 3: headers — "IP address" | "Business Justification"
    Data rows start at row 4.

    Operation (block vs allow) is inferred from the tab title row.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return

    # Infer operation from row 1 title
    title = str(rows[0][0] or "").lower()
    operation = "Block" if "block" in title else "Allow" if "allow" in title else "Block/Allow"

    # Find header row — must contain "ip address" as a standalone header cell,
    # not the title row which contains "ip address block or allow"
    header_idx = None
    col_ip = 0
    col_just = 1
    for i, row in enumerate(rows):
        row_lower = [str(c).lower().strip() if c else "" for c in row]
        # Exact match: cell IS "ip address" (not "ip address block or allow")
        if any(c == "ip address" for c in row_lower):
            header_idx = i
            for j, c in enumerate(row_lower):
                if "ip" in c and "address" in c:
                    col_ip = j
                elif "justif" in c or "business" in c:
                    col_just = j
            break

    start = (header_idx + 1) if header_idx is not None else 3
    for row in rows[start:]:
        ip = _cell(row, col_ip)
        just = _cell(row, col_just)
        if not ip:
            continue
        req.ip_block_allow.append(IPBlockAllow(
            ip_address=ip,
            operation=operation,
            justification=just,
        ))


# ---------------------------------------------------------------------------
# VPN tab
# ---------------------------------------------------------------------------

def _parse_vpn_tab(ws: Worksheet, req: FirewallRequest) -> None:
    """
    The VPN tab is a key-value layout with a local side (left) and Remote side (right).
    We scan for labelled rows and extract values from the adjacent cell(s).
    """
    vpn = VPNConfig()

    # Read all cells into a grid for flexible lookup
    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append([str(c).strip() if c is not None else "" for c in row])

    def _find_value(label_fragment: str, col_offset: int = 1, side_col: int | None = None) -> str:
        """Find a row containing label_fragment and return cell at col_offset from label cell."""
        for row in grid:
            for j, cell in enumerate(row):
                if label_fragment.lower() in cell.lower():
                    target_col = (side_col if side_col is not None else j) + col_offset
                    if 0 <= target_col < len(row):
                        return row[target_col].strip()
        return ""

    def _find_col(label_fragment: str) -> int:
        """Return column index of the first cell matching label_fragment."""
        for row in grid:
            for j, cell in enumerate(row):
                if label_fragment.lower() in cell.lower():
                    return j
        return 0

    # Determine which columns hold local vs Remote values
    # The local section header is typically around col B, Remote around col G+
    local_header_col = _find_col("Local Section")
    remote_header_col = _find_col("Remote Section")

    def _local_val(label: str) -> str:
        return _find_value(label, col_offset=1, side_col=local_header_col if local_header_col else None)

    def _remote_val(label: str) -> str:
        return _find_value(label, col_offset=1, side_col=remote_header_col if remote_header_col else None)

    # Local side
    vpn.local_device = _local_val("VPN Device")
    vpn.local_tunnel_endpoint = _local_val("VPN Tunnel Endpoint")
    vpn.local_ike_encryption = _local_val("IKE Encryption")
    vpn.local_auth_method = _local_val("Authentication Method")
    vpn.local_dh_group = _local_val("Diffie-Helman")
    vpn.local_sa_lifetime_p1 = _local_val("Security Association Lifetime")
    vpn.local_hash_p1 = _local_val("Hash")
    vpn.local_ipsec_encryption = _local_val("IPSEC Encryption")
    vpn.local_pfs = _local_val("Perfect Forward Secrecy")

    # Remote side
    vpn.remote_device = _remote_val("VPN Device")
    vpn.remote_tunnel_endpoint = _remote_val("VPN Tunnel Endpoint")
    vpn.remote_ike_encryption = _remote_val("IKE Encryption")
    vpn.remote_auth_method = _remote_val("Authentication Method")
    vpn.remote_dh_group = _remote_val("Diffie-Helman")
    vpn.remote_sa_lifetime_p1 = _remote_val("Security Association Lifetime")
    vpn.remote_hash_p1 = _remote_val("Hash")
    vpn.remote_ipsec_encryption = _remote_val("IPSEC Encryption")
    vpn.remote_pfs = _remote_val("Perfect Forward Secrecy")

    # Networks/Host Routes — collect all non-empty cells in rows labelled "Networks/Host Routes"
    for row in grid:
        row_str = " ".join(row).lower()
        if "networks/host routes" in row_str or "networks/hosts" in row_str:
            for j, cell in enumerate(row):
                if cell and "networks" not in cell.lower() and "hosts" not in cell.lower():
                    val = cell.strip()
                    if val:
                        if j <= (remote_header_col or 99) // 2:
                            vpn.local_networks.append(val)
                        else:
                            vpn.remote_networks.append(val)

    req.vpn_config = vpn


# ---------------------------------------------------------------------------
# Manual entry constructor
# ---------------------------------------------------------------------------

def parse_manual_entry(fields: dict[str, Any]) -> FirewallRequest:
    """
    Construct a FirewallRequest from a dictionary of manually entered fields.

    Expected keys (all optional, missing ones go into missing_fields):
      reference_id, rule_owner, business_contact, business_area,
      nuclear_environment, designation, estimated_completion,
      fw_rules: list of dicts with keys matching FWRule fields,
      group_changes: list of dicts matching GroupChange fields,
      ip_block_allow: list of dicts matching IPBlockAllow fields,
      vpn_config: dict matching VPNConfig fields

    Returns a FirewallRequest ready for analysis.
    """
    req = FirewallRequest(
        reference_id=str(fields.get("reference_id", "")),
        rule_owner=str(fields.get("rule_owner", "")),
        business_contact=str(fields.get("business_contact", "")),
        business_area=str(fields.get("business_area", "")),
        nuclear_environment=bool(fields.get("nuclear_environment", False)),
        designation=str(fields.get("designation", "")),
        estimated_completion=str(fields.get("estimated_completion", "")),
        source_file="manual_entry",
    )

    for rule_dict in fields.get("fw_rules", []):
        req.fw_rules.append(FWRule(
            rule_number=str(rule_dict.get("rule_number", "")),
            operation=str(rule_dict.get("operation", "Add")),
            source=str(rule_dict.get("source", "")),
            destination=str(rule_dict.get("destination", "")),
            port=str(rule_dict.get("port", "")),
            protocol=str(rule_dict.get("protocol", "N/A")),
            application=str(rule_dict.get("application", "")),
            justification=str(rule_dict.get("justification", "")),
        ))

    for grp in fields.get("group_changes", []):
        req.group_changes.append(GroupChange(
            group_name=str(grp.get("group_name", "")),
            new_membership=str(grp.get("new_membership", "")),
            action=str(grp.get("action", "")),
        ))

    for iba in fields.get("ip_block_allow", []):
        req.ip_block_allow.append(IPBlockAllow(
            ip_address=str(iba.get("ip_address", "")),
            operation=str(iba.get("operation", "Block")),
            justification=str(iba.get("justification", "")),
        ))

    if "vpn_config" in fields and fields["vpn_config"]:
        v = fields["vpn_config"]
        req.vpn_config = VPNConfig(
            local_device=str(v.get("local_device", "")),
            local_tunnel_endpoint=str(v.get("local_tunnel_endpoint", "")),
            local_ike_encryption=str(v.get("local_ike_encryption", "")),
            local_auth_method=str(v.get("local_auth_method", "")),
            local_dh_group=str(v.get("local_dh_group", "")),
            local_sa_lifetime_p1=str(v.get("local_sa_lifetime_p1", "")),
            local_hash_p1=str(v.get("local_hash_p1", "")),
            local_ipsec_encryption=str(v.get("local_ipsec_encryption", "")),
            local_pfs=str(v.get("local_pfs", "")),
            local_networks=list(v.get("local_networks", [])),
            remote_device=str(v.get("remote_device", "")),
            remote_tunnel_endpoint=str(v.get("remote_tunnel_endpoint", "")),
            remote_ike_encryption=str(v.get("remote_ike_encryption", "")),
            remote_auth_method=str(v.get("remote_auth_method", "")),
            remote_dh_group=str(v.get("remote_dh_group", "")),
            remote_sa_lifetime_p1=str(v.get("remote_sa_lifetime_p1", "")),
            remote_hash_p1=str(v.get("remote_hash_p1", "")),
            remote_ipsec_encryption=str(v.get("remote_ipsec_encryption", "")),
            remote_pfs=str(v.get("remote_pfs", "")),
            remote_networks=list(v.get("remote_networks", [])),
        )

    _validate(req)
    return req


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def _validate(req: FirewallRequest) -> None:
    """Populate req.missing_fields and req.warnings with actionable messages."""
    if not req.rule_owner:
        req.missing_fields.append("rule_owner")
    if not req.business_contact:
        req.missing_fields.append("business_contact")
    if not req.business_area:
        req.missing_fields.append("business_area")
    if not req.designation:
        req.missing_fields.append("designation")

    if not req.fw_rules and not req.group_changes and not req.ip_block_allow and not req.has_vpn():
        req.warnings.append(
            "No request content found — all tabs appear empty. "
            "Check that the spreadsheet is the correct template."
        )

    for rule in req.fw_rules:
        if rule.missing_fields:
            req.warnings.append(
                f"Rule {rule.rule_number or '?'} is missing: {', '.join(rule.missing_fields)}"
            )
        # Flag suspiciously broad rules
        if rule.source in ("any", "0.0.0.0", "0.0.0.0/0"):
            req.warnings.append(
                f"Rule {rule.rule_number or '?'}: source is '{rule.source}' — verify this is intentional"
            )
        if rule.destination in ("any", "0.0.0.0", "0.0.0.0/0"):
            req.warnings.append(
                f"Rule {rule.rule_number or '?'}: destination is '{rule.destination}' — verify this is intentional"
            )

    if req.nuclear_environment:
        req.warnings.append(
            "Request is flagged for Nuclear environment — regulatory documentation "
            "required (e.g. NERC CIP-005 Interactive Remote Access; "
            "substitute your own regime's equivalent if deploying elsewhere)"
        )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(row: tuple, col: int | None) -> str:
    if col is None or col >= len(row):
        return ""
    val = row[col]
    if val is None:
        return ""
    return str(val).strip()
